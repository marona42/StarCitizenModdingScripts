import os.path
import configparser
import openpyxl as xl  # pip install openpyxl needed!


class ConversionProject:
    def __init__(self, configname):
        if not os.path.isfile(configname):
            raise FileNotFoundError(
                f"essential config file '{configname}' Not found. aborting..."
            )
        self.refinidata = None
        self.filldata = {}
        self.newseg = [[[]], [[]]]
        self.modseg = [[[]], [[]]]
        self._load_config()
        self.refinidata = self._load_ini(self.refname)
        if self.fillmode:  # FIXME: not ini, xlsx!
            print(
                f"{self.__econfig['update']['dataxlsx']} file found! fill translated data..."
            )
            for tmp in self.tfilename:
                self.filldata.update(self._load_xlsx(tmp))
# self.filldata = self._load_xlsx(self.__econfig['update']['dataxlsx'])+  #load translated xlsx file
            print("translated data read done")
        if self.patchmode:
            print(
                f"{self.__econfig['update']['newrefini']} file found! use this file instead..."
            )
            self.prevrefdata = self.refinidata
            self.refinidata = self._load_ini(self.__econfig["update"]["newrefini"])
            print("new ref data read done")

            self.mnsegdat, self.phsegdat = {}, {}
            self.mnsegdat = self._load_ini(self._manualsegfname)
            self.phsegdat = self._load_ini(self._placeholdersegfname)

        # os.rename('global_ref.ini','global_ref_legacy.ini')
        # os.rename(self.__econfig['update']['newrefini'],'global_ref.ini')

    def write_xlsx(self, dot=100):
        print("write start")
        xlswb = [[xl.workbook.Workbook()], [xl.workbook.Workbook()]]  # primary, alt
        xlscnt = [1, 1]
        for wb in xlswb:
            wb[-1].active.append(
                ["context", "en", self.__econfig["convert"]["targetlang"], "comments"]
            )  # write initial first row

        for keyword in self.refinidata:
            if keyword.startswith(self.splitkeys):
                widx = 1
            else:
                widx = 0  # determine which document to write  (primary: 0/ alt: 1)

            if (
                self.refinidata[keyword] == ""
                or any(tmp in self.refinidata[keyword] for tmp in self.excludewords)
                or any(keyword.startswith(tmp) for tmp in self.excludesegs)
                or keyword.endswith(',P')       # New Placeholder format since 3.20 
            ):
                continue  # exclude including excludekeywords/segs and empty segments

            if self.patchmode:  # FIXME: 오차 있음
                if (
                    keyword in self.prevrefdata
                    and self.prevrefdata[keyword] != self.refinidata[keyword]
                ):
                    self.modseg[widx][-1].append(xlscnt[widx])
                if keyword not in self.prevrefdata:
                    self.newseg[widx][-1].append(xlscnt[widx])
                    # print(f"new seg {xlscnt[widx]} : {keyword}")
                    if keyword in self.mnsegdat:  # TODO: 로그파일로 출력
                        print(f"please check {keyword} at {self._manualsegfname}")
                    if keyword in self.phsegdat:
                        print(f"please check {keyword} at {self._placeholdersegfname}")

            if self.fillmode and keyword in self.filldata:
                if self.filldata[keyword].startswith("'"):
                    self.filldata[keyword] = (
                        "'" + self.filldata[keyword]
                    )  # '로 시작하는 거 하나 더 붙여주기
                if (
                    len(self.modseg[widx][-1]) != 0
                    and self.modseg[widx][-1][-1] == xlscnt[widx]
                ):
                    xlswb[widx][-1].active.append(
                        [
                            keyword,
                            self.refinidata[keyword],
                            self.filldata[keyword],
                            "modified",
                        ]
                    )
                else:
                    xlswb[widx][-1].active.append(
                        [keyword, self.refinidata[keyword], self.filldata[keyword]]
                    )
            else:
                if (
                    len(self.modseg[widx][-1]) != 0
                    and self.modseg[widx][-1][-1] == xlscnt[widx]
                ):
                    xlswb[widx][-1].active.append(
                        [keyword, self.refinidata[keyword], "", "modified"]
                    )
                elif (
                    len(self.newseg[widx][-1]) != 0
                    and self.newseg[widx][-1][-1] == xlscnt[widx]
                ):
                    xlswb[widx][-1].active.append(
                        [keyword, self.refinidata[keyword], "", "new"]
                    )
                else:
                    xlswb[widx][-1].active.append([keyword, self.refinidata[keyword]])
            xlscnt[widx] += 1

            if xlscnt[widx] > self.doculimit:
                xlswb[widx][-1].save(
                    filename=self.resname[widx] + f"_P{len(xlswb[widx])}.xlsx"
                )
                xlswb[widx].append(xl.workbook.Workbook())
                xlswb[widx][-1].append(
                    ["context", "en", self.__econfig["convert"]["targetlang"]]
                )
                self.modseg[widx].append([])
                self.newseg[widx].append([])
                xlscnt[widx] = 0

            if xlscnt[widx] % dot == 0:
                print(".", end="", flush=True)

        for i in range(2):
            xlswb[i][-1].save(filename=self.resname[i] + f"_P{len(xlswb[i])}.xlsx")
        print(f"\nwrite done xlsx {len(xlswb[0])}+{len(xlswb[1])} files.")
        if self.patchmode:
            print("write log files...")
            with open("segments_new.log", "w") as f:
                f.write("")
            with open("segments_modified.log", "w") as f:
                f.write("")

            self.write_info("segments_new", "\t\tnew segments", "")
            for docuidx in range(len(self.newseg[0])):
                self.write_info(
                    "segments_new",
                    f"\n\t{self.resname[0]}_P{docuidx}",
                    self.newseg[0][docuidx],
                    xlswb[0][docuidx],
                )
            for docuidx in range(len(self.newseg[1])):
                self.write_info(
                    "segments_new",
                    f"\n\t{self.resname[1]}_P{docuidx}",
                    self.newseg[1][docuidx],
                    xlswb[1][docuidx],
                )

            self.write_info("segments_modified", "\t\tmodified segments", "")
            for docuidx in range(len(self.modseg[0])):
                self.write_info(
                    "segments_modified",
                    f"\n\t{self.resname[0]}_P{docuidx}",
                    self.modseg[0][docuidx],
                    xlswb[0][docuidx],
                )
            for docuidx in range(len(self.modseg[1])):
                self.write_info(
                    "segments_modified",
                    f"\n\t{self.resname[1]}_P{docuidx}",
                    self.modseg[1][docuidx],
                    xlswb[1][docuidx],
                )

    def write_info(self, filename, title, dat, keyworddic=None):
        with open(filename + ".log", "a") as f:
            if dat == "":
                f.write(f"{title}\n")
                return
            else:
                f.write(f"{title} - {len(dat)} items\n")
            prevnum, isseq = 0, False
            for segnum in dat:
                if segnum - prevnum != 1:
                    if isseq:
                        f.write(f" - {prevnum}")
                    f.write(
                        f"\n{segnum}: ( {keyworddic.active['A'+str(segnum+1)].value} )"
                    )
                    isseq = False
                else:
                    isseq = True
                prevnum = segnum
            if isseq:
                f.write(f" - {prevnum}")
            f.write("\n\n")

    def _load_config(self):
        self.__econfig = configparser.ConfigParser(
            delimiters="=", strict=True, interpolation=None
        )
        self.__econfig.read("mconfig.ini")  # load settings ini file

        self.doculimit = int(self.__econfig["convert"]["splitlimit"])
        self.resname = [
            self.__econfig["convert"]["resname"],
            self.__econfig["convert"]["resaltname"],
        ]
        self.refname = self.__econfig["convert"]["refini"]

        self.tfilename = self.__econfig["update"]["dataxlsx"].split(",")
        self.fillmode = all(os.path.isfile(tmp) for tmp in self.tfilename)
        self.patchmode = os.path.isfile(self.__econfig["update"]["newrefini"])
        self._manualsegfname = self.__econfig["update"]["manualseg"]
        self._placeholdersegfname = self.__econfig["update"]["phseg"]

        self.excludewords = list(self.__econfig["parse"]["excludekeywords"].split(","))
        self.excludesegs = list(self.__econfig["parse"]["excludesegments"].split(","))
        self.splitkeys = tuple(self.__econfig["parse"]["splitsegment"].split(","))

    def _load_ini(self, filename):
        with open(filename, "r", encoding="utf​-8-sig") as f:
            tmp_str = "[DEFAULT]\n" + f.read()
        data = configparser.ConfigParser(
            delimiters="=", strict=True, interpolation=None
        )
        data.optionxform = str
        data.read_string(tmp_str)

        return data["DEFAULT"]

    def _load_xlsx(self, filename):
        res = dict()
        wd = xl.load_workbook(filename, read_only=True, data_only=True)
        ws = wd.active
        for rowd in ws.iter_rows(min_row=2):  # starts with 2nd row
            if rowd[2].value is not None:  # if only translated segment exist
                res[rowd[0].value] = rowd[2].value
        return res


if __name__ == "__main__":
    ConversionProject("mconfig.ini").write_xlsx()
    # sys.exit(main(sys.argv))
