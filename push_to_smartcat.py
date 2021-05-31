# push to smartcat project.
# first need pull from smartcat to make xlsx
# TODO: requires export-smartcat.py, smartcat.ini
import configparser
import openpyxl as xl     #pip install openpyxl needed!
#import export_smartcat

splitthreshold=50000

transdata={}

with open('global_ref.ini', 'r',encoding='utf​-8-sig') as f:        #target ref ini file
    origin_str = '[DEFAULT]\n' + f.read()
origindata = configparser.ConfigParser(delimiters='=',strict=True,interpolation=None)
origindata.optionxform=str
origindata.read_string(origin_str)

twb = xl.load_workbook("global_out.ini.xlsx")
tws = twb.active
for rowd in tws.iter_rows(min_row=2):
        #print(rowd[0].value)
        if len(rowd)>2: print(f"Warning: multiple rows detected in {rowd}")
        if rowd[1].value is not None:
            if '=' not in rowd[1].value: print(f"Warning: {rowd[1].value}")
            else:
                (inkey,indata) = rowd[1].value.split('=',1)
                transdata[inkey]=indata

partcnt,splitcnt=1,1
wb=xl.workbook.Workbook()
ws=wb.active
ws.append(["en","ko"])
for keyword in origindata['DEFAULT']:
    if ('(PH)' in origindata['DEFAULT'][keyword] or '[PH]' in origindata['DEFAULT'][keyword] or 'WIP' in origindata['DEFAULT'][keyword] or
        '*DELETE THIS*' in origindata['DEFAULT'][keyword] or 'DO NOT USE' in origindata['DEFAULT'][keyword] or 'PLACEHOLDER' in origindata['DEFAULT'][keyword] or
        origindata['DEFAULT'][keyword]=='' ): continue    #exclude by text
    if keyword in transdata:
        ws.append([f"{keyword}={origindata['DEFAULT'][keyword]}",f"{keyword}={transdata[keyword]}"])
    else:
        ws.append([f"{keyword}={origindata['DEFAULT'][keyword]}"])
    splitcnt+=1
    if splitcnt > splitthreshold:
        splitcnt=0
        wb.save(filename=f"global_push_P{partcnt}.xlsx")
        partcnt+=1
        wb=xl.workbook.Workbook()
        ws=wb.active
wb.save(filename=f"global_push_P{partcnt}.xlsx")
print(f"Split ini with {partcnt} xlsx file(s) Done")
