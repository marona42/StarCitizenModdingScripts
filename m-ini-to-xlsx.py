import sys
import os.path
import configparser
import openpyxl as xl     #pip install openpyxl needed!

class ConversionProject:
    def __init__(self,configname):
        if(not os.path.isfile(configname)): raise FileNotFoundError(f"essential config file '{configname}' Not found. aborting...")
        self.refinidata=None
        self.filldata=None
        self.newseg=[[[]],[[]]]
        self.modseg=[[[]],[[]]]
        self._load_config()
        self.refinidata=self._load_ini(self.refname)
        if self.fillmode :
            print(f"{self.__econfig['update']['dataini']} file found! fill translated data...")
            self.filldata = self._load_ini(self.__econfig['update']['dataini'])     #load translated ini file
            print("translated data read done")
        if self.patchmode:
            print(f"{self.__econfig['update']['newrefini']} file found! use this file instead...")
            self.prevrefdata = self.refinidata
            self.refinidata = self._load_ini(self.__econfig['update']['newrefini'])
            print("new ref data read done")

            self.mnsegdat,self.phsegdat = {},{}
            self.mnsegdat = self._load_ini(self._manualsegfname)
            self.phsegdat = self._load_ini(self._placeholdersegfname)

        #os.rename('global_ref.ini','global_ref_legacy.ini')
        #os.rename(self.__econfig['update']['newrefini'],'global_ref.ini')

    def write_xlsx(self,dot=100):
        print("write start")
        xlswb=[[xl.workbook.Workbook()],[xl.workbook.Workbook()]]   #primary, alt
        for wb in xlswb: wb[-1].active.append(["context","en",self.__econfig['convert']['targetlang']])   #write initial first row

        for keyword in self.refinidata:
            if any(keyword.startswith(tmp) for tmp in self.splitkeys ): widx=1
            else : widx=0  #determine which document to write  (primary: 0/ alt: 1)

            if any(tmp in self.refinidata[keyword] for tmp in self.excludewords) or self.refinidata[keyword]=='': continue    #exclude including excludekeywords and empty segments

            if self.patchmode:
                if keyword in self.prevrefdata and self.prevrefdata[keyword] != self.refinidata[keyword]:
                    self.modseg[widx][-1].append(xlswb[widx][-1].active.max_row) #FIXME: 어느 도큐의 번호인지?
                if keyword not in self.prevrefdata:
                    self.newseg[widx][-1].append(xlswb[widx][-1].active.max_row)
                    if keyword in self.mnsegdat: #TODO: 로그파일로 출력
                        print(f'please check {keyword} at {self._manualsegfname}')
                    if keyword in self.phsegdat:
                        print(f'please check {keyword} at {self._placeholdersegfname}')

            if self.fillmode and keyword in self.filldata:
                xlswb[widx][-1].active.append([keyword,self.refinidata[keyword],self.filldata[keyword]])
            else: 
                xlswb[widx][-1].active.append([keyword,self.refinidata[keyword]])

            if xlswb[widx][-1].active.max_row > self.doculimit:
                xlswb[widx][-1].save(filename=self.resname[widx]+f"_P{len(xlswb[widx])}.xlsx")
                xlswb[widx].append(xl.workbook.Workbook())
                self.modseg[widx][-1].append([])
                self.newseg[widx][-1].append([])

            if xlswb[widx][-1].active.max_row % dot == 0: print(".",end="",flush=True)

        for i in range(2): xlswb[i][-1].save(filename=self.resname[i]+f"_P{len(xlswb[i])}.xlsx")
        print(f"\nwrite done xlsx {len(xlswb[0])}+{len(xlswb[1])} files.")
        if self.patchmode:
            print("write log files...")
            self.write_info("segments_new","\t\tnew segments","")
            for docuidx in range(len(self.newseg[0])):
                self.write_info("segments_new",f"\n\t{self.resname[0]}_P{docuidx}",self.newseg[0][docuidx])
            for docuidx in range(len(self.newseg[1])):
                self.write_info("segments_new",f"\n\t{self.resname[1]}_P{docuidx}",self.newseg[1][docuidx])

            self.write_info("segments_modified","\t\tmodified segments","")
            for docuidx in range(len(self.modseg[0])):
                self.write_info("segments_modified",f"\n\t{self.resname[0]}_P{docuidx}",self.modseg[0][docuidx])
            for docuidx in range(len(self.modseg[1])):
                self.write_info("segments_modified",f"\n\t{self.resname[1]}_P{docuidx}",self.modseg[1][docuidx])

    def write_info(self,filename,title,dat):
        with open(filename+".log","a") as f:
            f.write(title+'\n')
            if dat == "": return
            prevnum,isseq=0,False
            for segnum in dat:
                if segnum-prevnum != 1:
                    if isseq:
                        f.write(f"-{prevnum}")
                    f.write(f"\n{segnum}")
                    isseq=False
                else:
                    isseq=True
                prevnum=segnum


    def _load_config(self):
        self.__econfig = configparser.ConfigParser(delimiters='=',strict=True,interpolation=None)
        self.__econfig.read('mconfig.ini')   #load settings ini file
        
        self.doculimit = int(self.__econfig['convert']['splitlimit'])
        self.resname   = [self.__econfig['convert']['resname'],self.__econfig['convert']['resaltname']]
        self.refname   = self.__econfig['convert']['refini']

        self.fillmode    = os.path.isfile(self.__econfig['update']['dataini'])
        self.patchmode = os.path.isfile(self.__econfig['update']['newrefini'])
        self._manualsegfname      = self.__econfig['update']['manualseg']
        self._placeholdersegfname = self.__econfig['update']['phseg']

        self.excludewords = list(self.__econfig['parse']['excludekeywords'].split(','))
        self.splitkeys    = list(self.__econfig['parse']['splitsegment'].split(','))

    def _load_ini(self,filename):
        with open(filename, 'r',encoding='utf​-8-sig') as f:
            tmp_str = '[DEFAULT]\n' + f.read()
        data = configparser.ConfigParser(delimiters='=',strict=True,interpolation=None)
        data.optionxform=str
        data.read_string(tmp_str)
        
        return data['DEFAULT']


if __name__ == "__main__":
    ConversionProject("mconfig.ini").write_xlsx()
    #sys.exit(main(sys.argv))
